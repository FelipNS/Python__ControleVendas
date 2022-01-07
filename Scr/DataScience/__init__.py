import json
from pymongo import MongoClient
import mysql.connector
import pandas as pd
import openpyxl as pyxl
import openpyxl.worksheet.table as pyxltb
from datetime import datetime

COLUMNS_NAME = ['Nº da Comanda',  'Vendedor', 'Combo', 'F. de Pagamento', 'Preço', 'Tamanho', 'Bairro', 'Cereais', 'Chocolates', 'Coberturas', 'Derivados do Leite', 'Diversos', 'Frutas', 'Obs', 'Data']
SELLER = ('seller', 'Vendedor')
COMBO = ('combos', 'Combo')
PAYMENT_OPTION = ('payment_option', 'F. de Pagamento')
SIZE = ('size', 'Tamanho')
NEIGHBORHOOD = ('neighborhood', 'Bairro')
DATE = ('date_time', 'Data')
class ManipulationDB:
    def __init__(self) -> None:
        """Create connection in databases
        """
        #Connect MongoDB
        self.client = MongoClient('localhost', 27017)
        self.db = self.client['acaiteria']

        #Connect MySQL
        self.conn = mysql.connector.connect(host='localhost', user='root', passwd='', database='acaiteria')
    
    def create_dataframe(self, name_or_id_seller='id') -> pd.DataFrame:
        """Create data frame with data of sells
        Args:
            name_or_id_seller (str, optional): Supports 'name' or 'id' If name, the seller column in dataframe will receive seller's name, if not, it will receive the seller id. Defaults to 'id'.
        Returns:
            pd.DataFrame
        """
        i = dict()
        collection = self.db['comandas']
        cursor = collection.find({})
        df_main = pd.DataFrame(self.__pretty_json(cursor[0], name_or_id_seller), columns=COLUMNS_NAME)
        for i in cursor[1::]:
            df_temp = pd.DataFrame(self.__pretty_json(i, name_or_id_seller), columns=COLUMNS_NAME)
            df_main = pd.concat([df_main, df_temp])
        
        cursor.close()
        return df_main
        
    def __pretty_json(self, sheet: str, name_or_id='id') -> dict:
        """Create a pretty json format
        Args:
            sheet (str): dict of a record
            name_or_id (str, optional): If 'name' the seller column will be seller name else will be ID seller. Defaults to id.
        Returns:
            dict: Dict containing record data 
        """
        old_json = sheet
        pretty_sheet = dict()
        pretty_sheet['number_sheet'] = old_json['n_comanda']
        pretty_sheet['seller'] = self._name_seller(old_json['seller']) if name_or_id  == 'name' else old_json['seller']
        pretty_sheet['combo'] = old_json['combo']
        pretty_sheet['payment_type'] = old_json['f_pagamento']
        pretty_sheet['price'] = old_json['preco']
        pretty_sheet['size'] = old_json['tamanho']
        pretty_sheet['neighborhood'] = old_json['bairro']
        pretty_sheet['cereals'] = list(i for i in old_json['cereais'])
        pretty_sheet['chocolates'] = list(i for i in old_json['chocolates'])
        pretty_sheet['covers'] = list(i for i in old_json['coberturas'])
        pretty_sheet['milk_derivatives'] = list(i for i in old_json['derivados do leite'])
        pretty_sheet['several'] = list(i for i in old_json['diversos'])
        pretty_sheet['fruits'] = list(i for i in old_json['frutas'])
        pretty_sheet['obs'] = old_json['obs']
        pretty_sheet['date'] = self._str_to_date(old_json['date_time'])
        json_pretty = json.dumps(pretty_sheet, indent=4, ensure_ascii=False).encode('utf8')
        json_pretty = [json.loads(json_pretty.decode()).values()]
        return json_pretty
    
    def statistical_equations(self) -> dict:
        """Calculate statistical metrics
        Returns:
            tuple: Contaning claculations results
        """
        df = self.create_dataframe()
        
        sales_quantity = df['Nº da Comanda'].count()
        revenue = df['Preço'].sum()
        mean_sales = round(df['Preço'].mean(), 2)
        return {"Quantity": sales_quantity, "Revenue": revenue, "Mean_Sales": mean_sales}
    
    def count_type(self, type: tuple)-> pd.DataFrame:
        """Create data frame group by any colum type
        Args:
            type (tuple): Value used to group the column. Suppot values: SELLER, COMBO, PAYMENT_OPTION, SIZE, NEIGHBORHOOD
        Returns:
            pd.DataFrame: Data frame group by column 
        """
        df = self.create_dataframe()
        quantitys = dict()
        if type != SELLER:
            cursor = self.db['dadosGerais'].find({'_id': type[0]})
            types_names = list(i for i in cursor[0].keys() if i != '_id')
        else:
            cursor = self.conn.cursor()
            cursor.execute('SELECT id FROM empregados')
            types_names = list(map(lambda x: x[0], cursor))
        if type == SIZE:
            types_names = list(map(self.__str_numeric, types_names))
        for i in types_names:
            qtd = 0
            for j in df[type[1]]:
                if j == i:
                    qtd += 1
            quantitys[i] = qtd 
        quantity_any_thing = pd.DataFrame.from_dict(quantitys, orient='index', columns=['Qtd'])
        cursor.close()
        final_df = quantity_any_thing.reset_index().copy()
        return final_df.sort_values(by='Qtd')
    
    def __str_numeric(self, str: str) -> int:
        """Return only number of a str
        Args:
            str (str): String to convert
        Returns:
            int: Only numbers os string
        """
        only_number = ''
        for i in str:
            if i.isnumeric():
                only_number += i
        if only_number == '1':
            only_number = '1000'
        return int(only_number)
    
    def _name_seller(self, id: int) -> str:
        """Return name of seller according to ID
        Args:
            id (int): ID seller
        Returns:
            str: Name seller
        """
        cursor = self.conn.cursor()
        cursor.execute(f'SELECT primeiro_nome, sobrenome FROM empregados WHERE id = {id}')
        employee_name = list(map(lambda x: f'{x[0]} {x[1]}', cursor))
        cursor.close()
        return employee_name[0]
    
    def _str_to_date(self, str: str):
        date_object = datetime.strptime(str, "%Y-%m-%d %H:%M:%S.%f")
        date = f'{date_object.day}/{date_object.month}/{date_object.year}'
        return date

class DataFrameToExcel:
    def __init__(self, dataframe: pd.DataFrame, path: str) -> None:
        self.df = dataframe
        self.manDB = ManipulationDB()
        self.path = path
        self.writer = pd.ExcelWriter(self.path, engine='xlsxwriter')

    def main_dataframe_to_excel(self) -> None:
        """Export data frame to excel file
        Args:
            dataframe (pd.DataFrame): Data frame to convert excel file
        """
        self.df.to_excel(self.writer, sheet_name='Todas as Comandas', encoding='utf-8')
    
    def auxiliary_dataframe(self) -> pd.DataFrame:
        df_sales_per_seller = self.manDB.count_type(SELLER)
        df_sales_per_combo = self.manDB.count_type(COMBO)
        df_sales_per_size = self.manDB.count_type(SIZE)
        df_sales_per_neighborhood = self.manDB.count_type(NEIGHBORHOOD)
        df_sales_per_payment_option = self.manDB.count_type(PAYMENT_OPTION)
        df_sales_per_day = self._date_dataframe()

        tpl_df = (df_sales_per_seller, df_sales_per_combo, df_sales_per_size, df_sales_per_neighborhood, df_sales_per_payment_option, df_sales_per_day)

        start_col = 0
        for i in tpl_df:
            i.to_excel(self.writer, sheet_name='Tabelas Auxiliares', encoding='utf-8', startcol= start_col)
            start_col += 4

    def _date_dataframe(self):
        all_dates = self.df['Data']
        unique_days = []
        for day in all_dates:
            if day not in unique_days:
                unique_days.append(day)
        sales_per_day = dict()
        for i in unique_days:
            qtd = 0
            for j in self.df['Data']:
                if i == j:
                    qtd += 1
            sales_per_day[i] = qtd    
        df_sales_per_day = pd.DataFrame.from_dict(sales_per_day, orient='index', columns=['Qtd'])
        df_sales_per_day['index'] = df_sales_per_day.index
        return df_sales_per_day[['index', 'Qtd']]
    
    def export_excel(self):
        self.main_dataframe_to_excel()
        self.auxiliary_dataframe()
        self.close_excel_file()

        workbook = pyxl.load_workbook(self.path)
        ws = workbook['Todas as Comandas']
        ws.delete_cols(1)
        ws = workbook['Tabelas Auxiliares']
        for i in (21, 17, 13, 9, 5, 1):
            ws.delete_cols(i)
        
        row = 2
        while ws[f'A{row}'].value != None:
            id = ws[f'A{row}'].value
            ws[f'A{row}'] = ManipulationDB()._name_seller(id)
            row += 1
            
        self._range_to_table(workbook)
        workbook.save(self.path)
    
    def _range_to_table(self, workbook):
        ws = workbook['Tabelas Auxiliares']

        for i in range(1,100):
            if ws.cell(1, i).value == None:
                if ws.cell(1, i+1).value == None:
                    break

        index_column = []
        for j in range(1, i):
            cell = ws.cell(1, j)
            if  cell.value == 'index':
                index_column.append(cell.column)

        for i in index_column:
            
            row = 1
            while ws.cell(row, i).value != None:
                row += 1
            tab = pyxltb.Table(displayName=f'Tb{index_column.index(i)}', ref=f"{ws.cell(1, i).column_letter}{1}:{ws.cell(row, i+1).column_letter}{row-1}")
            ws.add_table(tab)
        workbook.save(self.path)

    def close_excel_file(self):
        self.writer.close()
